# from io import BytesIO
#
# from django.contrib.auth.decorators import login_required
# from django.http import HttpResponse
# from django.shortcuts import render
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Font
#
# from accreditation.models import (
#     Course,
#     GraduationRequirement,
#     CourseIndicatorRelation,
# )
#
#
# def _lv_txt(v):
#     s = str(v or '').strip().lower()
#
#     if s in ['high', 'strong', '1', 'a', '强', '高']:
#         return '强'
#     if s in ['medium', 'middle', 'mid', '2', 'b', '中', '中等']:
#         return '中'
#     return '弱'
#
#
# def _lv_css(v):
#     s = str(v or '').strip().lower()
#
#     if s in ['high', 'strong', '1', 'a', '强', '高']:
#         return 'high'
#     if s in ['medium', 'middle', 'mid', '2', 'b', '中', '中等']:
#         return 'mid'
#     return 'low'
#
#
# def _make_rows(course_qs, req_qs):
#     rel_qs = CourseIndicatorRelation.objects.select_related('course', 'indicator').all()
#
#     rel_map = {}
#     for x in rel_qs:
#         rel_map[(x.course_id, x.indicator_id)] = x
#
#     all_rows = []
#
#     for c in course_qs:
#         one_row = []
#
#         for req in req_qs:
#             tmp = []
#             high_num = 0
#             mid_num = 0
#             low_num = 0
#
#             for indicator in req.indicators.all():
#                 rel = rel_map.get((c.id, indicator.id))
#                 if not rel:
#                     continue
#
#                 lv_txt = _lv_txt(rel.support_level)
#                 lv_css = _lv_css(rel.support_level)
#
#                 if lv_css == 'high':
#                     high_num += 1
#                 elif lv_css == 'mid':
#                     mid_num += 1
#                 else:
#                     low_num += 1
#
#                 tmp.append({
#                     'indicator_code': indicator.code,
#                     'level_text': lv_txt,
#                     'level_css': lv_css,
#                     'weight': rel.weight,
#                 })
#
#             one_row.append({
#                 'requirement': req,
#                 'items': tmp,
#                 'total_num': len(tmp),
#                 'high_num': high_num,
#                 'mid_num': mid_num,
#                 'low_num': low_num,
#             })
#
#         all_rows.append({
#             'course': c,
#             'cells': one_row,
#         })
#
#     return all_rows
#
#
# def _out_excel(req_qs, row_data):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = '支撑矩阵'
#
#     head = ['课程']
#     for req in req_qs:
#         head.append(req.code)
#     ws.append(head)
#
#     for row in row_data:
#         line = [f'{row["course"].code} {row["course"].name}']
#
#         for cell in row['cells']:
#             if not cell['items']:
#                 line.append('')
#                 continue
#
#             txt_arr = []
#             for item in cell['items']:
#                 if item['weight'] is not None and item['weight'] != '':
#                     txt_arr.append(f'{item["indicator_code"]}（{item["level_text"]}，{item["weight"]}）')
#                 else:
#                     txt_arr.append(f'{item["indicator_code"]}（{item["level_text"]}）')
#
#             line.append('；'.join(txt_arr))
#
#         ws.append(line)
#
#     for cell in ws[1]:
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#
#     for row in ws.iter_rows(min_row=2):
#         for cell in row:
#             cell.alignment = Alignment(vertical='top', wrap_text=True)
#
#     ws.column_dimensions['A'].width = 24
#
#     col_no = 2
#     for _ in req_qs:
#         if col_no <= 26:
#             ws.column_dimensions[chr(64 + col_no)].width = 24
#         col_no += 1
#
#     fp = BytesIO()
#     wb.save(fp)
#     fp.seek(0)
#
#     resp = HttpResponse(
#         fp.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     resp['Content-Disposition'] = 'attachment; filename="support_matrix.xlsx"'
#     return resp
#
#
# # 支撑矩阵页
# @login_required
# def support_matrix(request):
#     course_id = request.GET.get('course_id', '').strip()
#     requirement_id = request.GET.get('requirement_id', '').strip()
#     out_type = request.GET.get('export', '').strip()
#
#     course_qs = Course.objects.all().order_by('code', 'id')
#     req_qs = GraduationRequirement.objects.prefetch_related('indicators').all().order_by('code', 'id')
#
#     if course_id:
#         course_qs = course_qs.filter(id=course_id)
#
#     if requirement_id:
#         req_qs = req_qs.filter(id=requirement_id)
#
#     row_data = _make_rows(course_qs, req_qs)
#
#     if out_type == 'excel':
#         return _out_excel(req_qs, row_data)
#
#     context = {
#         'course_rows': Course.objects.all().order_by('code', 'id'),
#         'requirement_rows': GraduationRequirement.objects.all().order_by('code', 'id'),
#         'course_id': course_id,
#         'requirement_id': requirement_id,
#         'requirements': req_qs,
#         'matrix_rows': row_data,
#     }
#     return render(request, 'accreditation/support_matrix.html', context)
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from accreditation.models import (
    Course,
    GraduationRequirement,
    CourseIndicatorRelation,
)


# 支撑强度文本
def _lv_txt(v):
    s = str(v or '').strip().lower()

    if s in ['high', 'strong', '1', 'a', '强', '高']:
        return '强'
    if s in ['medium', 'middle', 'mid', '2', 'b', '中', '中等']:
        return '中'
    return '弱'


# 支撑强度样式标记
def _lv_css(v):
    s = str(v or '').strip().lower()

    if s in ['high', 'strong', '1', 'a', '强', '高']:
        return 'high'
    if s in ['medium', 'middle', 'mid', '2', 'b', '中', '中等']:
        return 'mid'
    return 'low'


# 组装矩阵行数据
def _make_rows(course_qs, req_list):
    rel_qs = CourseIndicatorRelation.objects.select_related('course', 'indicator').all()

    rel_map = {}
    for x in rel_qs:
        rel_map[(x.course_id, x.indicator_id)] = x

    all_rows = []

    for c in course_qs:
        one_row = []
        row_total_num = 0
        row_high_num = 0
        row_mid_num = 0
        row_low_num = 0

        for req in req_list:
            tmp = []
            high_num = 0
            mid_num = 0
            low_num = 0

            for indicator in req.indicators.all():
                rel = rel_map.get((c.id, indicator.id))
                if not rel:
                    continue

                lv_txt = _lv_txt(rel.support_level)
                lv_css = _lv_css(rel.support_level)

                if lv_css == 'high':
                    high_num += 1
                elif lv_css == 'mid':
                    mid_num += 1
                else:
                    low_num += 1

                tmp.append({
                    'indicator_code': indicator.code,
                    'level_text': lv_txt,
                    'level_css': lv_css,
                    'weight': rel.weight,
                })

            total_num = len(tmp)
            row_total_num += total_num
            row_high_num += high_num
            row_mid_num += mid_num
            row_low_num += low_num

            one_row.append({
                'requirement': req,
                'items': tmp,
                'total_num': total_num,
                'high_num': high_num,
                'mid_num': mid_num,
                'low_num': low_num,
            })

        all_rows.append({
            'course': c,
            'cells': one_row,
            'total_num': row_total_num,
            'high_num': row_high_num,
            'mid_num': row_mid_num,
            'low_num': row_low_num,
        })

    return all_rows


# 组装毕业要求表头汇总
def _make_requirement_summary(req_list, row_data):
    summary = []

    for idx, req in enumerate(req_list):
        total_num = 0
        high_num = 0
        mid_num = 0
        low_num = 0
        coverage_num = 0

        for row in row_data:
            cell = row['cells'][idx]
            if cell['total_num']:
                coverage_num += 1

            total_num += cell['total_num']
            high_num += cell['high_num']
            mid_num += cell['mid_num']
            low_num += cell['low_num']

        summary.append({
            'id': req.id,
            'code': req.code,
            'name': req.name,
            'total_num': total_num,
            'high_num': high_num,
            'mid_num': mid_num,
            'low_num': low_num,
            'coverage_num': coverage_num,
        })

    return summary


# 页面顶部统计
def _make_stats(row_data, req_list):
    relation_total = 0
    high_total = 0
    mid_total = 0
    low_total = 0
    active_cell_num = 0

    for row in row_data:
        relation_total += row['total_num']
        high_total += row['high_num']
        mid_total += row['mid_num']
        low_total += row['low_num']

        for cell in row['cells']:
            if cell['total_num']:
                active_cell_num += 1

    return {
        'course_num': len(row_data),
        'requirement_num': len(req_list),
        'relation_total': relation_total,
        'high_total': high_total,
        'mid_total': mid_total,
        'low_total': low_total,
        'active_cell_num': active_cell_num,
    }


# 导出 Excel
def _out_excel(req_list, row_data):
    wb = Workbook()
    ws = wb.active
    ws.title = '支撑矩阵'

    head = ['课程']
    for req in req_list:
        head.append(req.code)
    ws.append(head)

    for row in row_data:
        line = [f'{row["course"].code} {row["course"].name}']

        for cell in row['cells']:
            if not cell['items']:
                line.append('')
                continue

            txt_arr = []
            for item in cell['items']:
                if item['weight'] is not None and item['weight'] != '':
                    txt_arr.append(f'{item["indicator_code"]}（{item["level_text"]}，{item["weight"]}）')
                else:
                    txt_arr.append(f'{item["indicator_code"]}（{item["level_text"]}）')

            line.append('；'.join(txt_arr))

        ws.append(line)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.column_dimensions['A'].width = 24

    col_no = 2
    for _ in req_list:
        if col_no <= 26:
            ws.column_dimensions[chr(64 + col_no)].width = 24
        col_no += 1

    fp = BytesIO()
    wb.save(fp)
    fp.seek(0)

    resp = HttpResponse(
        fp.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="support_matrix.xlsx"'
    return resp


# 支撑矩阵页
@login_required
def support_matrix(request):
    course_id = request.GET.get('course_id', '').strip()
    requirement_id = request.GET.get('requirement_id', '').strip()
    out_type = request.GET.get('export', '').strip()

    course_qs = Course.objects.all().order_by('code', 'id')
    req_qs = GraduationRequirement.objects.prefetch_related('indicators').all().order_by('code', 'id')

    if course_id:
        course_qs = course_qs.filter(id=course_id)

    if requirement_id:
        req_qs = req_qs.filter(id=requirement_id)

    req_list = list(req_qs)
    row_data = _make_rows(course_qs, req_list)
    requirement_summary = _make_requirement_summary(req_list, row_data)
    stats = _make_stats(row_data, req_list)

    if out_type == 'excel':
        return _out_excel(req_list, row_data)

    context = {
        'course_rows': Course.objects.all().order_by('code', 'id'),
        'requirement_rows': GraduationRequirement.objects.all().order_by('code', 'id'),
        'course_id': course_id,
        'requirement_id': requirement_id,
        'requirements': req_list,
        'requirement_summary': requirement_summary,
        'matrix_rows': row_data,
        'stats': stats,
    }
    return render(request, 'accreditation/support_matrix.html', context)